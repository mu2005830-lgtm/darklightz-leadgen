"""
Export leads to Excel (.xlsx) — Darklightz Lead Generator v2.0.

v2.0: All new fields added (email, mobile, category, maps_link,
latitude, longitude, opening_hours).
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Sequence

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# (header label, dict key, column width)
COLUMNS = [
    ("Business Name",   "name",          32),
    ("Category",        "category",      20),
    ("Phone",           "phone",         18),
    ("Mobile",          "mobile",        18),
    ("WhatsApp",        "whatsapp",      34),
    ("Email",           "email",         30),
    ("Instagram",       "instagram",     32),
    ("Facebook",        "facebook",      32),
    ("Website",         "website",       32),
    ("Address",         "address",       38),
    ("Google Maps",     "maps_link",     40),
    ("Latitude",        "latitude",      14),
    ("Longitude",       "longitude",     14),
    ("Rating",          "rating",         8),
    ("Reviews",         "reviews",       10),
    ("Opening Hours",   "opening_hours", 40),
    ("Lead Type",       "lead_type",     18),
    ("Status",          "status",        16),
    ("City",            "city",          14),
    ("Keyword",         "keyword",       18),
    ("Notes",           "notes",         50),
    ("Date Added",      "created_at",    22),
]

HEADER_BG  = "1E1E2E"
HEADER_FG  = "CDD6F4"
ALT_ROW_BG = "181825"
NORMAL_BG  = "1E1E2E"
ACCENT_FG  = "CBA6F7"


def export(leads: Sequence[dict], output_path: str) -> str:
    """Write *leads* to *output_path* (.xlsx).  Returns absolute path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    thin   = Side(style="thin", color="313244")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Header row --------------------------------------------------
    header_font  = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    header_fill  = PatternFill("solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # ---- Data rows ---------------------------------------------------
    data_align  = Alignment(vertical="top", wrap_text=True)
    normal_font = Font(name="Calibri", size=10, color="CDD6F4")
    link_font   = Font(name="Calibri", size=10, color="89B4FA", underline="single")

    # Columns that contain URLs — render as hyperlinks
    url_key_indices = {
        key: idx + 1
        for idx, (_, key, _) in enumerate(COLUMNS)
        if key in ("whatsapp", "instagram", "facebook", "website", "maps_link")
    }

    for row_idx, lead in enumerate(leads, start=2):
        bg       = ALT_ROW_BG if row_idx % 2 == 0 else NORMAL_BG
        row_fill = PatternFill("solid", fgColor=bg)

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = str(lead.get(key, "") or "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.alignment = data_align
            cell.border    = border

            if col_idx in url_key_indices.values() and value.startswith("http"):
                cell.font      = link_font
                cell.hyperlink = value
            else:
                cell.font = normal_font

    # ---- Sheet styling -----------------------------------------------
    ws.sheet_view.showGridLines = False
    wb.properties.title   = "Darklightz Lead Generator — Export"
    wb.properties.creator = "Darklightz Studio"
    wb.properties.created = datetime.now()

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return os.path.abspath(output_path)
